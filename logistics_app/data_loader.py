"""
data_loader.py — Lee tarifas y catálogo de producto desde Excel.
Exporta cargar_todo() → dict DATOS que sustituye a las constantes de config.py.
"""
from __future__ import annotations

import re
from pathlib import Path
import openpyxl

# ── Rutas por defecto (relativas a logistics_app → un nivel arriba) ────────────
_APP_DIR  = Path(__file__).parent
BASE_DIR  = _APP_DIR.parent

DEFAULT_TARIFAS_PATH  = BASE_DIR / "tarifas_logisticas.xlsx"
DEFAULT_CATALOGO_PATH = BASE_DIR / "catalogo_productos.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES INTERNAS
# ══════════════════════════════════════════════════════════════════════════════

def _safe_float(val, default=None):
    """Convierte val a float; devuelve default si falla."""
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _safe_int(val, default=None):
    """Convierte val a int; devuelve default si falla."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _extract_first_int(text: str) -> int | None:
    """Extrae el primer entero de un string ('8 palés' → 8)."""
    m = re.search(r'\d[\d,\.]*', str(text))
    if not m:
        return None
    return int(re.sub(r'[,\.]', '', m.group()).split()[0])


def _extract_last_int(text: str) -> int | None:
    """Extrae el último número entero de un string.
    Maneja separadores de miles: '3.500 cajas' → 3500, '5,000 kg' → 5000.
    """
    # Buscar secuencias de dígitos con posibles separadores de miles (. o ,)
    nums = re.findall(r'\d[\d\.,]*\d|\d', str(text))
    if not nums:
        return None
    raw = nums[-1]
    # Eliminar separadores de miles (. o ,) — asumimos que son miles, no decimales
    cleaned = raw.replace('.', '').replace(',', '')
    try:
        return int(cleaned)
    except ValueError:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# LECTURA DE TARIFAS
# ══════════════════════════════════════════════════════════════════════════════

def cargar_tarifas(path=None) -> dict:
    """
    Lee tarifas_logisticas.xlsx y devuelve un dict con:
      - transporte_peso : list[(kg_max, precio_peninsula, precio_baleares|None)]
      - baleares_kg_max : float  (máximo kg con tabla Baleares)
      - tarifa_pale_provincia : dict{PROVINCIA_UPPER: precio_€}
      - cargas_completas : dict{PROVINCIA_UPPER: {n_pales: {kg_max, precio}}}
      - almacen_madrid : dict{clave: valor}
      - almacen_regional_cajas : list[(cajas_max, €_por_caja)]
      - almacen_regional_min_baleares : float
      - almacen_regional_recepcion_caja : float
      - almacen_regional_manipulacion_pedido : float
      - almacen_regional_seguro_pct : float
    """
    if path is None:
        path = DEFAULT_TARIFAS_PATH
    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(
            f"No se encuentra el archivo de tarifas: {path}\n"
            "Asegúrate de que 'tarifas_logisticas.xlsx' está en la carpeta LOGÍSTICA."
        )

    wb = openpyxl.load_workbook(path, data_only=True)
    datos = {}

    # ── 1. Transporte_Kg ──────────────────────────────────────────────────────
    ws = wb["Transporte_Kg"]

    peninsula_dict = {}   # kg → precio
    baleares_dict  = {}   # kg → precio

    for row in ws.iter_rows(min_row=4, values_only=True):
        # Península: cols A (0), B (1)
        kg_p = _safe_float(row[0] if len(row) > 0 else None)
        pr_p = _safe_float(row[1] if len(row) > 1 else None)
        if kg_p is not None and pr_p is not None:
            peninsula_dict[kg_p] = pr_p

        # Baleares: cols D (3), E (4)
        kg_b = _safe_float(row[3] if len(row) > 3 else None)
        pr_b = _safe_float(row[4] if len(row) > 4 else None)
        if kg_b is not None and pr_b is not None:
            baleares_dict[kg_b] = pr_b

    baleares_kg_max = max(baleares_dict.keys()) if baleares_dict else 200.0

    # Combinar: (kg, precio_peninsula, precio_baleares|None)
    transporte_peso = []
    for kg in sorted(peninsula_dict.keys()):
        transporte_peso.append((kg, peninsula_dict[kg], baleares_dict.get(kg, None)))

    datos["transporte_peso"]   = transporte_peso
    datos["baleares_kg_max"]   = baleares_kg_max

    # ── 2. Pale_Provincia ─────────────────────────────────────────────────────
    ws = wb["Pale_Provincia"]
    tarifa_pale = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        # Par izquierdo: A (0), B (1)
        prov1 = row[0] if len(row) > 0 else None
        prec1 = _safe_float(row[1] if len(row) > 1 else None)
        if prov1 is not None and prec1 is not None:
            tarifa_pale[str(prov1).strip().upper()] = prec1

        # Par derecho: D (3), E (4)
        prov2 = row[3] if len(row) > 3 else None
        prec2 = _safe_float(row[4] if len(row) > 4 else None)
        if prov2 is not None and prec2 is not None:
            tarifa_pale[str(prov2).strip().upper()] = prec2

    # Media peninsular como fallback
    if tarifa_pale:
        tarifa_pale["PENINSULA_MEDIA"] = round(
            sum(tarifa_pale.values()) / len(tarifa_pale), 2
        )

    datos["tarifa_pale_provincia"] = tarifa_pale

    # ── 3. Cargas_Completas ───────────────────────────────────────────────────
    ws = wb["Cargas_Completas"]
    all_rows = list(ws.iter_rows(min_row=3, values_only=True))

    # Fila 3 (idx 0): cabecera de palés → "8 palés", "10 palés"...
    # Fila 4 (idx 1): cabecera de kg   → "≤ 5.000 kg"...
    # Filas 5+ (idx 2+): datos por provincia

    pales_header = all_rows[0] if len(all_rows) > 0 else []
    kg_header    = all_rows[1] if len(all_rows) > 1 else []

    # Mapear col_index → (num_pales, kg_max)
    col_map = {}  # col_idx → {"pales": int, "kg_max": int}
    for col_i, val in enumerate(pales_header):
        if col_i == 0:
            continue  # skip province name col
        n_pales = _extract_first_int(str(val)) if val is not None else None
        if n_pales is None:
            continue
        kg_raw = kg_header[col_i] if col_i < len(kg_header) else None
        n_kg   = _safe_int(_extract_last_int(str(kg_raw))) if kg_raw is not None else None
        if n_kg is not None:
            col_map[col_i] = {"pales": n_pales, "kg_max": n_kg}

    cargas_completas = {}
    for row in all_rows[2:]:  # desde fila 5 en Excel
        if not row or row[0] is None:
            continue
        prov = str(row[0]).strip().upper()
        if not prov or prov.startswith("PALÉS") or prov.startswith("KG"):
            continue
        prov_data = {}
        for col_i, info in col_map.items():
            if col_i < len(row):
                precio = _safe_float(row[col_i])
                if precio is not None:
                    prov_data[info["pales"]] = {
                        "kg_max": info["kg_max"],
                        "precio": precio,
                    }
        cargas_completas[prov] = prov_data

    datos["cargas_completas"] = cargas_completas

    # ── 4. Almacen_Madrid ─────────────────────────────────────────────────────
    # Filas 4-16, col A=concepto, col B=unidad, col C=precio
    # Posición fija (i = row - 4):
    #  0→almacenaje_m3_dia, 1→recepcion_kg, 2→manipulados_pedido,
    #  3→manipulados_unidad_bulto, 4→gestion_admin_pedido, 5→inventarios_hora,
    #  6→desmanipulados_pedido, 7→desmanipulados_unidad, 8→sms,
    #  9→acondicionamiento_botella, 10→seguro, 11→recogidas_unidad, 12→duas_unidad
    ws = wb["Almacen_Madrid"]
    MADRID_KEYS = [
        "almacenaje_m3_dia",
        "recepcion_kg",
        "manipulados_pedido",
        "manipulados_unidad_bulto",
        "gestion_admin_pedido",
        "inventarios_hora",
        "desmanipulados_pedido",
        "desmanipulados_unidad",
        "sms",
        "acondicionamiento_botella",
        "seguro",
        "recogidas_unidad",
        "duas_unidad",
    ]

    almacen_madrid = {}
    madrid_rows = list(ws.iter_rows(min_row=4, max_row=16, values_only=True))
    for i, key in enumerate(MADRID_KEYS):
        if i < len(madrid_rows):
            row = madrid_rows[i]
            precio = _safe_float(row[2] if len(row) > 2 else None)
            if precio is not None:
                almacen_madrid[key] = precio

    datos["almacen_madrid"] = almacen_madrid

    # ── 5. Almacen_Regional ───────────────────────────────────────────────────
    ws = wb["Almacen_Regional"]

    # Tramos de cajas: filas 5-9 (col A=descripción, col B=unidad, col C=precio)
    # Extraemos el máximo de cajas del texto descriptivo
    almacen_regional_cajas = []
    tramo_rows = list(ws.iter_rows(min_row=5, max_row=9, values_only=True))
    for row in tramo_rows:
        desc  = str(row[0]) if row[0] is not None else ""
        prec  = _safe_float(row[2] if len(row) > 2 else None)
        cajas = _extract_last_int(desc)
        if cajas is not None and prec is not None:
            almacen_regional_cajas.append((cajas, prec))

    datos["almacen_regional_cajas"] = almacen_regional_cajas

    # Fila 10: mínimo Baleares (col C)
    row10 = list(ws.iter_rows(min_row=10, max_row=10, values_only=True))[0]
    datos["almacen_regional_min_baleares"] = _safe_float(
        row10[2] if len(row10) > 2 else None, default=170.0
    )

    # Fila 14: recepción/caja, fila 15: manipulación/pedido, fila 16: seguro
    otras_rows = list(ws.iter_rows(min_row=14, max_row=16, values_only=True))
    datos["almacen_regional_recepcion_caja"]      = _safe_float(
        otras_rows[0][2] if otras_rows else None, default=0.63
    )
    datos["almacen_regional_manipulacion_pedido"] = _safe_float(
        otras_rows[1][2] if len(otras_rows) > 1 else None, default=8.20
    )
    datos["almacen_regional_seguro_pct"]          = _safe_float(
        otras_rows[2][2] if len(otras_rows) > 2 else None, default=0.0008
    )

    wb.close()
    return datos


# ══════════════════════════════════════════════════════════════════════════════
# LECTURA DE PRODUCTOS
# ══════════════════════════════════════════════════════════════════════════════

def cargar_productos(path=None) -> dict:
    """
    Lee catalogo_productos.xlsx y devuelve dict {nombre: {dims, peso, ...}}
    con sólo los productos activos.
    Columnas del Excel (fila 4+):
      A=ID, B=Nombre, C=Fecha_Alta, D=Fecha_Baja, E=Uds_por_Caja,
      F=Largo_cm, G=Ancho_cm, H=Alto_cm, I=Peso_kg, J=Valor_Caja_EUR
    """
    if path is None:
        path = DEFAULT_CATALOGO_PATH
    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(
            f"No se encuentra el catálogo de productos: {path}\n"
            "Asegúrate de que 'catalogo_productos.xlsx' está en la carpeta LOGÍSTICA."
        )

    from datetime import date as _date, datetime as _datetime
    today = _datetime.today().date()

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    productos = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        pid = str(row[0]).strip()
        if pid.startswith("←") or pid == "":
            continue

        nombre = row[1]
        if nombre is None:
            continue
        nombre = str(nombre).strip()
        if nombre.startswith("←") or nombre == "":
            continue

        # Filtrar productos de baja
        fecha_baja = row[3]  # col D
        if fecha_baja is not None:
            if isinstance(fecha_baja, (_date, _datetime)):
                baja_date = fecha_baja.date() if isinstance(fecha_baja, _datetime) else fecha_baja
                if baja_date <= today:
                    continue  # producto dado de baja

        try:
            productos[nombre] = {
                "uds_por_caja": _safe_int(row[4],  default=1),    # E
                "largo_cm":     _safe_float(row[5], default=40.0), # F
                "ancho_cm":     _safe_float(row[6], default=30.0), # G
                "alto_cm":      _safe_float(row[7], default=25.0), # H
                "peso_kg":      _safe_float(row[8], default=5.0),  # I
                "valor_caja":   _safe_float(row[9], default=50.0), # J
            }
        except Exception:
            pass  # fila malformada, ignorar

    wb.close()
    return productos


# ══════════════════════════════════════════════════════════════════════════════
# PUNTO DE ENTRADA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def cargar_todo(tarifas_path=None, catalogo_path=None) -> dict:
    """
    Carga tarifas + catálogo y devuelve el dict DATOS completo.
    Lanza FileNotFoundError o ValueError con mensaje claro si algo falla.
    """
    datos = cargar_tarifas(tarifas_path)
    datos["productos"] = cargar_productos(catalogo_path)
    return datos


# ══════════════════════════════════════════════════════════════════════════════
# GENERADORES DE PLANTILLAS XLSX
# ══════════════════════════════════════════════════════════════════════════════

def _hdr_cell(ws, row, col, value, bg="1E3A5F"):
    """Escribe una celda de cabecera con estilo."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    return c


def _data_cell(ws, row, col, value, bold=False, bg=None):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name="Arial", size=9)
    c.alignment = Alignment(horizontal="left" if isinstance(value, str) else "center",
                            vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


def _provincias_ordenadas(datos: dict) -> list[str]:
    """Devuelve las 47 provincias en orden alfabético (sin PENINSULA_MEDIA)."""
    return sorted(
        p for p in datos.get("tarifa_pale_provincia", {}).keys()
        if p != "PENINSULA_MEDIA"
    )


def generar_plantilla_stock(datos: dict) -> bytes:
    """
    Plantilla de stock: filas = Central Madrid + 47 provincias,
    columnas = ALMACÉN + un producto por columna.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    productos   = list(datos.get("productos", {}).keys())
    almacenes   = ["Central Madrid"] + _provincias_ordenadas(datos)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    # Cabecera
    _hdr_cell(ws, 1, 1, "ALMACÉN", bg="1E3A5F")
    for j, prod in enumerate(productos, 2):
        _hdr_cell(ws, 1, j, prod, bg="2C5282")

    # Filas de datos
    for i, alm in enumerate(almacenes, 2):
        bg_row = "D6E4F0" if alm == "Central Madrid" else ("F7FAFC" if i % 2 == 0 else None)
        _data_cell(ws, i, 1, alm, bold=(alm == "Central Madrid"), bg=bg_row)
        for j in range(len(productos)):
            _data_cell(ws, i, 2 + j, 0, bg=bg_row)

    # Anchos
    ws.column_dimensions["A"].width = 24
    for j in range(len(productos)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 16
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generar_plantilla_llegadas(datos: dict) -> bytes:
    """
    Plantilla de llegadas: filas = Central Madrid + 47 provincias,
    columnas = ALMACÉN | FECHA | un producto por columna.
    Rellena la FECHA cuando lleguen mercancías a ese almacén.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    productos = list(datos.get("productos", {}).keys())
    almacenes = ["Central Madrid"] + _provincias_ordenadas(datos)

    wb = Workbook()
    ws = wb.active
    ws.title = "Llegadas"

    # Cabecera
    _hdr_cell(ws, 1, 1, "ALMACÉN",  bg="1E6B3C")
    _hdr_cell(ws, 1, 2, "FECHA",    bg="1E6B3C")
    for j, prod in enumerate(productos, 3):
        _hdr_cell(ws, 1, j, prod, bg="276749")

    # Filas de datos
    for i, alm in enumerate(almacenes, 2):
        bg_row = "D5F5E3" if alm == "Central Madrid" else ("F0FFF4" if i % 2 == 0 else None)
        _data_cell(ws, i, 1, alm,  bold=(alm == "Central Madrid"), bg=bg_row)
        _data_cell(ws, i, 2, None, bg=bg_row)   # FECHA en blanco
        for j in range(len(productos)):
            _data_cell(ws, i, 3 + j, 0, bg=bg_row)

    # Anchos
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    for j in range(len(productos)):
        ws.column_dimensions[get_column_letter(3 + j)].width = 16
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generar_plantilla_envios(datos: dict) -> bytes:
    """
    Plantilla de envíos planificados: filas = 47 provincias,
    columnas = PROVINCIA | ZONA | FECHA | un producto por columna.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    productos   = list(datos.get("productos", {}).keys())
    provincias  = _provincias_ordenadas(datos)

    wb = Workbook()
    ws = wb.active
    ws.title = "Envíos"

    # Cabecera
    _hdr_cell(ws, 1, 1, "PROVINCIA", bg="C0392B")
    _hdr_cell(ws, 1, 2, "ZONA",      bg="C0392B")
    _hdr_cell(ws, 1, 3, "FECHA",     bg="C0392B")
    for j, prod in enumerate(productos, 4):
        _hdr_cell(ws, 1, j, prod, bg="922B21")

    # Filas de datos
    for i, prov in enumerate(provincias, 2):
        bg_row = "FADBD8" if i % 2 == 0 else None
        _data_cell(ws, i, 1, prov,        bold=True, bg=bg_row)
        _data_cell(ws, i, 2, "peninsula", bg=bg_row)
        _data_cell(ws, i, 3, None,        bg=bg_row)  # FECHA en blanco
        for j in range(len(productos)):
            _data_cell(ws, i, 4 + j, 0, bg=bg_row)

    # Anchos
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    for j in range(len(productos)):
        ws.column_dimensions[get_column_letter(4 + j)].width = 16
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Test rápido al ejecutar directamente ──────────────────────────────────────
if __name__ == "__main__":
    import json, sys

    print("Cargando tarifas desde:", DEFAULT_TARIFAS_PATH)
    print("Cargando catálogo desde:", DEFAULT_CATALOGO_PATH)
    try:
        d = cargar_todo()
        print(f"\n✅ OK")
        print(f"   Tramos transporte peninsular : {len(d['transporte_peso'])}")
        print(f"   Tramos Baleares (hasta)      : {d['baleares_kg_max']} kg")
        print(f"   Provincias palé              : {len(d['tarifa_pale_provincia'])}")
        print(f"   Provincias cargas completas  : {len(d['cargas_completas'])}")
        print(f"   Conceptos almacén Madrid     : {len(d['almacen_madrid'])}")
        print(f"   Tramos almacén regional      : {len(d['almacen_regional_cajas'])}")
        print(f"   Productos activos            : {len(d['productos'])}")

        # Detalle de un par de valores
        print(f"\n   Almacenaje Madrid (€/m³/día) : {d['almacen_madrid'].get('almacenaje_m3_dia')}")
        print(f"   Recepción regional (€/caja)  : {d['almacen_regional_recepcion_caja']}")
        print(f"   Tarifa palé MADRID           : {d['tarifa_pale_provincia'].get('MADRID')}")
        print(f"   Cargas GUADALAJARA 8 palés   : {d['cargas_completas'].get('GUADALAJARA', {}).get(8)}")
    except Exception as e:
        print(f"\n❌ Error: {e}", file=sys.stderr)
        sys.exit(1)
